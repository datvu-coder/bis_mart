"""Import data from Bismart_standardized.xlsx into the backend database."""
from __future__ import annotations
import os, sys, sqlite3, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl first"); sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "bismart.db"
EXCEL = BASE_DIR.parent / "Bismart_standardized.xlsx"

def dt_to_str(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, datetime.date): return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time): return v.strftime("%H:%M:%S")
    return str(v) if v else None

def safe_int(v, default=0):
    if v is None: return default
    try: return int(v)
    except: return default

def safe_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

def rows_of(ws, start=2):
    """Yield non-empty rows as list of values starting from row `start`."""
    for row in ws.iter_rows(min_row=start, max_col=ws.max_column, values_only=True):
        if all(v is None for v in row):
            continue
        yield list(row)

def main():
    if not EXCEL.exists():
        print(f"Excel not found: {EXCEL}"); sys.exit(1)

    # Delete old DB to start fresh
    if DB_PATH.exists():
        DB_PATH.unlink()
        print(f"Deleted old {DB_PATH}")

    # Create schema
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    schema = (BASE_DIR / "schema.sql").read_text(encoding="utf-8")
    conn.executescript(schema)
    conn.commit()

    wb = openpyxl.load_workbook(str(EXCEL), data_only=True)
    stats = {}

    # ===== 1. EMPLOYEES (Sheet: "Nhân viên") =====
    ws = wb["Nhân viên"]
    emp_code_to_id = {}
    count = 0
    for r in rows_of(ws):
        # Cols: ID, Họ và tên, Mã NV, Ngày sinh, CCCD, Nơi ở, Tình trạng, Chức vụ,
        #       Phòng ban, Cửa hàng, Tỉnh, Khu vực, Ngày tạo, Ngày thử việc, Ngày HĐ,
        #       Ngày nghỉ, Lý do nghỉ, SĐT, Email, Mật khẩu, Ảnh, Mã CH, Vị trí, Điểm, Cấp bậc
        code = str(r[2]).strip() if r[2] else None
        if not code: continue
        if code in emp_code_to_id: continue  # skip duplicates
        name = str(r[1] or "").strip()
        position = str(r[7] or "PG").strip()
        pos_map = {"PG": "PG", "TLD": "TLD", "ADM": "ADM", "MNG": "MNG", "CS": "CS",
                   "ASM": "ASM", "TMK": "TMK"}
        position = pos_map.get(position, position)
        store = str(r[9] or "").strip()
        province = str(r[10] or "").strip() if r[10] else None
        area = str(r[11] or "").strip() if r[11] else None
        
        is_active = 1 if not r[15] else 0
        
        cur = conn.execute(
            "INSERT INTO employees (full_name, employee_code, date_of_birth, cccd, address, "
            "status, position, department, work_location, province, area, created_date, "
            "probation_date, official_date, resign_date, resign_reason, phone, email, "
            "password, avatar_url, store_code, geo_position, score, rank_level, is_active) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (name, code, dt_to_str(r[3]), str(r[4]) if r[4] else None,
             str(r[5]) if r[5] else None, str(r[6] or "Chính thức"),
             position, str(r[8] or "Kinh doanh"), store, province, area,
             dt_to_str(r[12]), dt_to_str(r[13]), dt_to_str(r[14]),
             dt_to_str(r[15]), str(r[16]) if r[16] else None,
             str(r[17]) if r[17] else None, str(r[18]) if r[18] else None,
             str(r[19]) if r[19] else "1111", str(r[20]) if r[20] else None,
             str(r[21]) if r[21] else None, str(r[22]) if r[22] else None,
             safe_int(r[23]), str(r[24]) if r[24] else None, is_active)
        )
        emp_code_to_id[code] = cur.lastrowid
        count += 1
    conn.commit()
    stats["Nhân viên"] = count

    # Create admin user from first MNG employee
    from werkzeug.security import generate_password_hash
    admin_emp_id = None
    for code, eid in emp_code_to_id.items():
        row = conn.execute("SELECT position FROM employees WHERE id=?", (eid,)).fetchone()
        if row and row[0] == "MNG":
            admin_emp_id = eid
            break
    if not admin_emp_id and emp_code_to_id:
        admin_emp_id = list(emp_code_to_id.values())[0]
    
    conn.execute(
        "INSERT INTO users (username, password_hash, employee_id) VALUES (?, ?, ?)",
        ("admin", generate_password_hash("admin123", method="pbkdf2:sha256"), admin_emp_id)
    )
    conn.commit()

    # ===== 2. PERMISSIONS (Sheet: "Phân quyền") =====
    ws = wb["Phân quyền"]
    count = 0
    for r in rows_of(ws):
        pos = str(r[1] or "").strip()
        if not pos: continue
        x = lambda v: 1 if str(v or "").strip().upper() == "X" else 0
        conn.execute(
            "INSERT OR IGNORE INTO permissions (position, description, can_attendance, can_report, "
            "can_manage_attendance, can_employees, can_more, can_crud, can_switch_store, "
            "can_store_list, can_product_list) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (pos, str(r[2] or ""), x(r[3]), x(r[4]), x(r[5]), x(r[6]),
             x(r[7]), x(r[8]), x(r[9]), x(r[10]), x(r[11]))
        )
        count += 1
    conn.commit()
    stats["Phân quyền"] = count

    # ===== 3. STORES (Sheet: "Danh sách cửa hàng") =====
    ws = wb["Danh sách cửa hàng"]
    store_code_to_id = {}
    count = 0
    for r in rows_of(ws):
        coords = str(r[0] or "").strip() if r[0] else ""
        lat, lon = None, None
        if coords and "," in coords:
            parts = coords.split(",")
            try:
                lat = float(parts[0].strip())
                lon = float(parts[1].strip())
            except: pass
        name = str(r[1] or "").strip()
        scode = str(r[3] or "").strip()
        if not name or not scode: continue
        try:
            cur = conn.execute(
                "INSERT INTO stores (name, store_code, store_group, latitude, longitude, "
                "province, sup, status, open_date, close_date, store_type, address, phone, owner, tax_code) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (name, scode, str(r[2] or "I"), lat, lon,
                 str(r[4] or "") if r[4] else None, str(r[5] or "") if r[5] else None,
                 str(r[6] or "Hoạt động"), dt_to_str(r[7]), dt_to_str(r[8]),
                 str(r[9] or "") if r[9] else None, str(r[10] or "") if r[10] else None,
                 str(r[11] or "") if r[11] else None, str(r[12] or "") if r[12] else None,
                 str(r[13] or "") if r[13] else None)
            )
            store_code_to_id[scode] = cur.lastrowid
            count += 1
        except sqlite3.IntegrityError:
            pass
    conn.commit()
    stats["Cửa hàng"] = count

    # ===== 4. STORE MANAGERS (Sheet: "Danh sách quản lý") =====
    ws = wb["Danh sách quản lý"]
    count = 0
    for r in rows_of(ws):
        store_name = str(r[1] or "").strip()
        emp_code = str(r[3] or "").strip()
        if not store_name or not emp_code: continue
        # Find store_id by name
        row = conn.execute("SELECT id FROM stores WHERE name=?", (store_name,)).fetchone()
        store_id = row[0] if row else None
        emp_id = emp_code_to_id.get(emp_code)
        if store_id and emp_id:
            try:
                conn.execute("INSERT INTO store_managers (store_id, employee_id) VALUES (?,?)",
                             (store_id, emp_id))
                count += 1
            except sqlite3.IntegrityError:
                pass
    conn.commit()
    stats["Quản lý"] = count

    # ===== 5. PRODUCTS (Sheet: "Danh sách sản phẩm") =====
    ws = wb["Danh sách sản phẩm"]
    count = 0
    for r in rows_of(ws):
        name = str(r[1] or "").strip()
        if not name: continue
        conn.execute(
            "INSERT INTO products (name, unit, price_with_vat, product_condition, product_group) "
            "VALUES (?,?,?,?,?)",
            (name, str(r[2] or "LON"), safe_float(r[3]),
             str(r[4] or "") if r[4] else None, str(r[5] or "DELI"))
        )
        count += 1
    conn.commit()
    stats["Sản phẩm"] = count

    # ===== 6. WORK SHIFTS (Sheet: "Danh sách ca") =====
    ws = wb["Danh sách ca"]
    count = 0
    for r in rows_of(ws):
        eid = str(r[0] or "").strip() if r[0] else None
        name = str(r[1] or "").strip()
        code = str(r[2] or "").strip() if r[2] else None
        if not name: continue
        sh, sm, eh, em_ = 0, 0, 0, 0
        if isinstance(r[3], datetime.time):
            sh, sm = r[3].hour, r[3].minute
        if isinstance(r[4], datetime.time):
            eh, em_ = r[4].hour, r[4].minute
        store = str(r[5] or "").strip() if r[5] else None
        conn.execute(
            "INSERT INTO work_shifts (excel_id, name, shift_code, start_hour, start_minute, "
            "end_hour, end_minute, store_name) VALUES (?,?,?,?,?,?,?,?)",
            (eid, name, code, sh, sm, eh, em_, store)
        )
        count += 1
    conn.commit()
    stats["Ca làm việc"] = count

    # ===== 7. ATTENDANCE (Sheet: "Chấm công") =====
    ws = wb["Chấm công"]
    count = 0
    for r in rows_of(ws):
        emp_code = str(r[2] or "").strip()
        emp_id = emp_code_to_id.get(emp_code)
        if not emp_id: continue
        attend_date = dt_to_str(r[6]) if r[6] else None
        if not attend_date: continue
        # Only take YYYY-MM-DD part
        if len(attend_date) > 10:
            attend_date = attend_date[:10]
        
        conn.execute(
            "INSERT INTO attendances (excel_id, employee_id, attend_date, shift_name, "
            "shift_time_range, coordinates, distance_in, check_in_time, check_in_diff, "
            "check_in_status, distance_out, check_out_time, check_out_diff, check_out_status) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (str(r[0] or ""), emp_id, attend_date, str(r[4] or ""),
             str(r[5] or ""), str(r[7] or ""), safe_float(r[8]),
             dt_to_str(r[9]), safe_int(r[10], None), str(r[11] or ""),
             safe_float(r[12]) if r[12] else None,
             dt_to_str(r[13]), safe_int(r[14], None) if r[14] else None,
             str(r[15] or "") if r[15] else None)
        )
        count += 1
    conn.commit()
    stats["Chấm công"] = count

    # ===== 8. SALES REPORTS (Sheet: "Báo cáo") =====
    ws = wb["Báo cáo"]
    excel_report_id_to_db = {}
    count = 0
    for r in rows_of(ws):
        eid = str(r[0] or "").strip()
        if not eid: continue
        cur = conn.execute(
            "INSERT INTO sales_reports (excel_id, report_date, pg_name, store_name, nu, "
            "sale_out, store_code, report_month, revenue, points, employee_code) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (eid, dt_to_str(r[1]) or "", str(r[2] or ""), str(r[3] or ""),
             safe_int(r[4]), safe_float(r[5]), str(r[6] or ""),
             safe_int(r[7]), safe_float(r[8]), safe_int(r[9]),
             str(r[10] or ""))
        )
        excel_report_id_to_db[eid] = cur.lastrowid
        count += 1
    conn.commit()
    stats["Báo cáo"] = count

    # ===== 9. SALE ITEMS (Sheet: "Báo cáo chi tiết") =====
    ws = wb["Báo cáo chi tiết"]
    count = 0
    batch = []
    for r in rows_of(ws):
        item_eid = str(r[0] or "").strip()
        report_eid = str(r[1] or "").strip()
        db_report_id = excel_report_id_to_db.get(report_eid)
        if not db_report_id: continue
        batch.append((
            item_eid, db_report_id, report_eid, None,
            str(r[5] or ""), str(r[6] or ""), safe_int(r[7]),
            safe_float(r[8]), str(r[9] or ""), str(r[10] or "")
        ))
        count += 1
        if len(batch) >= 500:
            conn.executemany(
                "INSERT INTO sale_items (excel_id, report_id, report_excel_id, product_id, "
                "product_name, unit, quantity, unit_price, product_group, store_code) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)", batch)
            batch = []
    if batch:
        conn.executemany(
            "INSERT INTO sale_items (excel_id, report_id, report_excel_id, product_id, "
            "product_name, unit, quantity, unit_price, product_group, store_code) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)", batch)
    conn.commit()
    stats["Báo cáo chi tiết"] = count

    # ===== 10. COURSE TITLES (Sheet: "Tiêu đề") =====
    ws = wb["Tiêu đề"]
    count = 0
    for r in rows_of(ws):
        eid = str(r[0] or "").strip()
        if not eid: continue
        conn.execute(
            "INSERT OR IGNORE INTO course_titles (excel_id, title, access_level, image_url, "
            "description, rating, target_group) VALUES (?,?,?,?,?,?,?)",
            (eid, str(r[1] or ""), str(r[2] or ""), str(r[3] or "") if r[3] else None,
             str(r[4] or "") if r[4] else None, safe_float(r[5]) if r[5] else None,
             str(r[6] or "") if r[6] else None)
        )
        count += 1
    conn.commit()
    stats["Tiêu đề (khóa học)"] = count

    # ===== 11. COURSE CONTENTS (Sheet: "Nội dung") =====
    ws = wb["Nội dung"]
    count = 0
    for r in rows_of(ws):
        eid = str(r[0] or "").strip()
        if not eid: continue
        conn.execute(
            "INSERT OR IGNORE INTO course_contents (excel_id, title_id, title, detail_html, "
            "points, attachment_type, image_url, video_url, file_url, embed_code, status) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (eid, str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
             safe_int(r[4]), str(r[5] or "") if r[5] else None,
             str(r[6] or "") if r[6] else None, str(r[7] or "") if r[7] else None,
             str(r[8] or "") if r[8] else None, str(r[9] or "") if r[9] else None,
             str(r[10] or "Đang kiểm tra"))
        )
        count += 1
    conn.commit()
    stats["Nội dung (bài học)"] = count

    # ===== 12. COURSE ENROLLMENTS (Sheet: "Lịch sử tham gia") =====
    ws = wb["Lịch sử tham gia"]
    count = 0
    for r in rows_of(ws):
        conn.execute(
            "INSERT INTO course_enrollments (excel_id, title_id, employee_code, full_name, enrolled_at) "
            "VALUES (?,?,?,?,?)",
            (str(r[0] or ""), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
             dt_to_str(r[4]) if r[4] else str(r[4] or ""))
        )
        count += 1
    conn.commit()
    stats["Lịch sử tham gia"] = count

    # ===== 13. COURSE COMPLETIONS (Sheet: "Lịch sử hoàn thành") =====
    ws = wb["Lịch sử hoàn thành"]
    count = 0
    for r in rows_of(ws):
        conn.execute(
            "INSERT INTO course_completions (excel_id, title_id, content_id, employee_code, "
            "full_name, completed_at, points, content_name) VALUES (?,?,?,?,?,?,?,?)",
            (str(r[0] or ""), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
             str(r[4] or ""), dt_to_str(r[5]) if r[5] else str(r[5] or ""),
             safe_int(r[6]), str(r[7] or ""))
        )
        count += 1
    conn.commit()
    stats["Lịch sử hoàn thành"] = count

    # ===== 14. QUIZ QUESTIONS (Sheet: "Câu hỏi") =====
    ws = wb["Câu hỏi"]
    count = 0
    for r in rows_of(ws):
        q = str(r[1] or "").strip()
        if not q: continue
        conn.execute(
            "INSERT INTO quiz_questions (question_type, question, option_a, option_b, "
            "option_c, option_d, correct_answer, points, content_id, question_number) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (str(r[0] or "TN"), q, str(r[2] or "") if r[2] else None,
             str(r[3] or "") if r[3] else None, str(r[4] or "") if r[4] else None,
             str(r[5] or "") if r[5] else None, str(r[6] or ""),
             safe_int(r[7]), str(r[8] or "") if r[8] else None, safe_int(r[9]))
        )
        count += 1
    conn.commit()
    stats["Câu hỏi"] = count

    # ===== 15. QUIZ RESULTS (Sheet: "Kết quả") =====
    ws = wb["Kết quả"]
    count = 0
    for r in rows_of(ws):
        conn.execute(
            "INSERT INTO quiz_results (submitted_at, employee_code, full_name, store_name, "
            "phone, content_id, score, answers_json) VALUES (?,?,?,?,?,?,?,?)",
            (dt_to_str(r[0]), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
             str(r[4] or ""), str(r[5] or ""), str(r[6] or ""), str(r[7] or ""))
        )
        count += 1
    conn.commit()
    stats["Kết quả"] = count

    # ===== 16. CLASS SCHEDULES (Sheet: "Lịch học") =====
    ws = wb["Lịch học"]
    count = 0
    for r in rows_of(ws):
        eid = str(r[0] or "").strip()
        if not eid: continue
        conn.execute(
            "INSERT OR IGNORE INTO class_schedules (excel_id, start_date, start_time, "
            "end_date, end_time, content, link, attendance_file) VALUES (?,?,?,?,?,?,?,?)",
            (eid, dt_to_str(r[1]), dt_to_str(r[2]), dt_to_str(r[3]),
             dt_to_str(r[4]), str(r[5] or ""), str(r[6] or "") if r[6] else None,
             str(r[7] or "") if r[7] else None)
        )
        count += 1
    conn.commit()
    stats["Lịch học"] = count

    # ===== 17. CLASS ATTENDANCES (Sheet: "Điểm danh lớp") =====
    ws = wb["Điểm danh lớp"]
    count = 0
    for r in rows_of(ws):
        conn.execute(
            "INSERT INTO class_attendances (schedule_id, attendance_id, employee_code, "
            "full_name, store_name, content, action, attend_time, attend_date) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (str(r[0] or ""), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
             str(r[4] or ""), str(r[5] or ""), str(r[6] or ""),
             dt_to_str(r[7]), dt_to_str(r[8]) if r[8] else str(r[8] or ""))
        )
        count += 1
    conn.commit()
    stats["Điểm danh lớp"] = count

    # ===== 18. AI TOOLS (Sheet: "AI") =====
    ws = wb["AI"]
    count = 0
    for r in rows_of(ws):
        conn.execute("INSERT INTO ai_tools (name, link) VALUES (?,?)",
                     (str(r[1] or ""), str(r[2] or "") if r[2] else None))
        count += 1
    conn.commit()
    stats["AI"] = count

    # ===== 19. AI USAGE (Sheet: "Lịch sử sử dụng AI") =====
    ws = wb["Lịch sử sử dụng AI"]
    count = 0
    for r in rows_of(ws):
        conn.execute(
            "INSERT INTO ai_usage_logs (excel_id, employee_code, full_name, store_name, "
            "ai_name, used_at, points) VALUES (?,?,?,?,?,?,?)",
            (str(r[0] or ""), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
             str(r[4] or ""), dt_to_str(r[5]), safe_int(r[6]))
        )
        count += 1
    conn.commit()
    stats["Lịch sử AI"] = count

    # ===== 20. LIKE HISTORY (Sheet: "Lịch sử like") =====
    ws = wb["Lịch sử like"]
    count = 0
    for r in rows_of(ws):
        conn.execute(
            "INSERT INTO post_likes (excel_id, ref_id, employee_code, full_name, points, created_at) "
            "VALUES (?,?,?,?,?,?)",
            (str(r[0] or ""), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""),
             safe_int(r[4]), dt_to_str(r[5]) or "")
        )
        count += 1
    conn.commit()
    stats["Lịch sử like"] = count

    conn.close()

    print("\n=== IMPORT COMPLETE ===")
    total = 0
    for name, cnt in stats.items():
        print(f"  {name}: {cnt} records")
        total += cnt
    print(f"\n  TOTAL: {total} records imported")
    print(f"  Database: {DB_PATH}")

if __name__ == "__main__":
    main()
